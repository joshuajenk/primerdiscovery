from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# export to excel file
def export_top_primers_to_excel(df, full_sequence, output_folder="output"):    #output dir
    output_dir = Path(output_folder)
    output_dir.mkdir(parents=True, exist_ok=True)

    # file attr
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"primers_{timestamp}.xlsx"
    output_path = output_dir / filename

    # worbook setup
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All Primers"

    # append data
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)

    # highlight warnings
    warning_fill = PatternFill(start_color="FFFACC", end_color="FFFACC", fill_type="solid")
    for row in ws1.iter_rows(min_row=2, min_col=7, max_col=7):  # Assuming col 7 = "Warnings"
        for cell in row:
            if cell.value:
                cell.fill = warning_fill

    # quality scores chart
    chart = BarChart()
    chart.title = "Primer Quality Scores"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Sequence"

    data_col = 3  # qs column
    label_col = 2  # seq column

    # categorization in excel, add chart based on QC
    data = Reference(ws1, min_col=data_col, min_row=1, max_row=len(df)+1)
    cats = Reference(ws1, min_col=label_col, min_row=2, max_row=len(df)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws1.add_chart(chart, "I2")

    wb.save(output_path)
    print(f"Full primer list exported to: {output_path}")